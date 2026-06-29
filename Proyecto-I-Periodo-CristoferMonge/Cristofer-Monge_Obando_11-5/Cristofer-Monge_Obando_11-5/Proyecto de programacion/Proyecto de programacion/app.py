# -*- coding: utf-8 -*-
"""
Sistema de Gestión de Contactos - DataConnect
nombre: Cristofer Monge Obando
seccion: 11-5,
proyecto de: Programación para Web 
"""

import os
import re
from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook

# =========================================================================
# CONFIGURACIÓN INICIAL DE LA APLICACIÓN
# =========================================================================

app = Flask(__name__)
app.secret_key = 'clave_secreta_dataconnect'  # Esta clave es para proteger las sesiones

# Aquí defino el nombre del archivo Excel que va a funcionar como mi "base de datos"
EXCEL_FILE = 'contactos.xlsx'

# Esta expresión regular la uso para validar correos electrónicos tanto al crear como al editar
# Básicamente asegura que tenga el formato: algo@algo.dominio (con 2 a 4 letras en el dominio)
EMAIL_REGEX = r'^[^\s@]+@[^\s@]+\.[a-zA-Z]{2,4}$'

# =========================================================================
# FUNCIÓN PARA INICIALIZAR EL ARCHIVO EXCEL (Indicador 13)
# =========================================================================

def inicializar_excel():
    """
    Esta función se encarga de crear el archivo Excel con los encabezados que necesito
    si es que no existe todavía. Es como crear una base de datos desde cero.
    Los encabezados son los que me pidieron en la Sección 1 del proyecto.
    """
    # Verifico si el archivo ya existe en el sistema
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()  # Creo un libro de trabajo nuevo
        ws = wb.active    # Tomo la hoja activa
        ws.title = "Contactos"  # Le pongo un nombre descriptivo
        
        # Aquí van los encabezados obligatorios que pide el proyecto
        ws.append(["Nombre", "Apellido", "Teléfono", "Correo", "Dirección", "Categoría", "Favorito"])
        wb.save(EXCEL_FILE)  # Guardo el archivo en el sistema

# Llamo a la función para que se ejecute apenas arranca la aplicación
inicializar_excel()

# =========================================================================
# CREDENCIALES DE ACCESO (Indicador 3.1)
# =========================================================================

# Estas son las credenciales fijas que uso para el login
USUARIO_CORRECTO = 'admin'    # El usuario que debe escribir el usuario
CONTRASENA_CORRECTA = '1234'  # La contraseña que debe escribir

# =========================================================================
# RUTA PRINCIPAL - REDIRECCIÓN AL LOGIN
# =========================================================================

@app.route('/')
def inicio():
    """
    Cuando alguien entra a la raíz de mi aplicación, lo redirijo automáticamente
    a la pantalla de login. Así me aseguro que nadie pase sin autenticarse.
    """
    return redirect(url_for('login'))

# =========================================================================
# RUTA DE LOGIN - AUTENTICACIÓN DE USUARIOS
# =========================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Esta función controla el acceso al sistema. Cuando el usuario envía el
    formulario de login, verifico que sus credenciales coincidan con las que
    tengo configuradas. Si son correctas, lo mando a la pantalla principal.
    Si no, le muestro un mensaje de error.
    """
    error = None  # Variable para guardar mensajes de error
    
    if request.method == 'POST':  # Si el usuario envió el formulario
        usuario = request.form.get('usuario')      # Obtengo el usuario escrito
        contrasena = request.form.get('contrasena')  # Obtengo la contraseña escrita
        
        # Verifico si las credenciales coinciden con las que tengo guardadas
        if usuario == USUARIO_CORRECTO and contrasena == CONTRASENA_CORRECTA:
            return redirect(url_for('principal'))  # Si es correcto, lo dejo pasar
        else:
            error = 'Usuario o contraseña incorrectos'  # Si no, muestro error
    
    # Muestro el formulario de login (con el error si lo hay)
    return render_template('login.html', error=error)

# =========================================================================
# RUTA PRINCIPAL - LISTA DE CONTACTOS CON BÚSQUEDA Y ORDENAMIENTO
# =========================================================================

@app.route('/principal')
def principal():
    """
    Esta es la función más importante. Carga todos los contactos desde el Excel,
    aplica filtros de búsqueda si el usuario buscó algo, y ordena alfabéticamente
    si el usuario pidió ordenarlos. Cumple con el Indicador 6 (Ordenamiento) y
    el Indicador 7 (Búsqueda).
    """
    # Me aseguro que el archivo Excel exista
    inicializar_excel()
    
    # Cargo el archivo Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Capturo los parámetros que el usuario envió desde la URL (por GET)
    query = request.args.get('q', '').strip().lower()  # Término de búsqueda
    ordenar_alfa = request.args.get('ordenar', '')     # ¿Ordenar alfabéticamente?

    lista_contactos = []  # Aquí voy a guardar todos los contactos
    
    # Recorro todas las filas del Excel, saltándome la primera (que son los encabezados)
    # El enumerate me da el número de fila para usar como ID
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None:  # Solo proceso si hay un nombre en la fila
            
            # -------------------- LIMPIEZA DE DATOS --------------------
            # Convierto el nombre a string, manejando el caso de números
            val_nombre = row[0]
            if isinstance(val_nombre, float) and val_nombre.is_integer():
                nombre = str(int(val_nombre)).strip()
            else:
                nombre = str(val_nombre).strip()

            # Hago lo mismo con el apellido
            val_apellido = row[1]
            if isinstance(val_apellido, float) and val_apellido.is_integer():
                apellido = str(int(val_apellido)).strip()
            else:
                apellido = str(val_apellido).strip() if row[1] else ''

            # Los demás campos los convierto a string directamente
            telefono = str(row[2]).strip() if row[2] else ''
            correo = str(row[3]).strip() if row[3] else ''
            direccion = str(row[4]).strip() if row[4] else ''
            categoria = str(row[5]).strip() if row[5] else 'Otro'
            favorito = str(row[6]).strip() if row[6] else 'No'
            
            # Creo el nombre completo en minúsculas para facilitar la búsqueda
            nombre_completo_virtual = f"{nombre} {apellido}".lower().strip()
            
            # -------------------- MOTOR DE BÚSQUEDA (Indicador 7) --------------------
            # Si el usuario escribió algo en la búsqueda, filtro los contactos
            if query:
                # Primero verifico si es una coincidencia exacta por nombre o nombre completo
                if query == nombre.lower() or query == nombre_completo_virtual:
                    encontrado = True
                else:
                    # Si es un solo dígito, busco en nombre y apellido
                    if len(query) == 1 and query.isdigit():
                        encontrado = (query in nombre.lower() or query in apellido.lower())
                    else:
                        # Búsqueda parcial en todos los campos relevantes
                        encontrado = (query in nombre_completo_virtual or 
                                      query in nombre.lower() or 
                                      query in apellido.lower() or 
                                      query in telefono or 
                                      query in correo.lower() or
                                      query in direccion.lower())
                
                # Si no encontré coincidencia, salto a la siguiente fila
                if not encontrado:
                    continue

            # Creo el diccionario con todos los datos del contacto
            contacto = {
                'id': index,                    # El número de fila en Excel
                'nombre': nombre,
                'apellido': apellido,
                'telefono': telefono,
                'correo': correo,
                'direccion': direccion,
                'categoria': categoria,
                'favorito': favorito
            }
            lista_contactos.append(contacto)  # Agrego el contacto a mi lista
    
    # -------------------- ORDENAMIENTO ALFABÉTICO (Indicador 6) --------------------
    # Si el usuario pidió ordenar, uso sorted() para ordenar por nombre
    if ordenar_alfa == 'si':
        lista_contactos = sorted(lista_contactos, key=lambda c: c['nombre'].lower())
    
    # Renderizo la página principal con la lista de contactos
    return render_template('index.html', contactos=lista_contactos, query_actual=request.args.get('q', ''))

# =========================================================================
# RUTA PARA CREAR CONTACTOS - FORMULARIO
# =========================================================================

@app.route('/crear_contacto')
def Crear_contactos():
    """
    Esta ruta solo muestra el formulario para crear nuevos contactos.
    Es como decirle al usuario: "Aquí tienes el formulario, llénalo".
    """
    return render_template('Crear_contactos.html')

# =========================================================================
# RUTA PARA GUARDAR CONTACTOS - PROCESAMIENTO Y VALIDACIÓN
# =========================================================================

@app.route('/guardar_contacto', methods=['POST'])
def guardar_contacto():
    """
    Esta es la función que procesa el formulario cuando el usuario crea un contacto.
    Valido todo en el backend (seguridad extra) y si todo está bien, lo guardo en Excel.
    Cumple con el Indicador 5 (Validaciones de Backend) y el Indicador 13 (Estructura).
    """
    # Obtengo todos los datos del formulario
    nombre_completo = request.form.get('nombre_completo', '').strip()
    telefono = request.form.get('telefono', '').strip()
    correo = request.form.get('correo', '').strip()
    categoria = request.form.get('categoria', 'Otro')
    direccion = request.form.get('direccion', '').strip()

    # -------------------- VALIDACIÓN 1: CAMPOS OBLIGATORIOS --------------------
    # Verifico que los campos más importantes no estén vacíos
    if not nombre_completo or not telefono or not correo:
        print("Error de Backend: Campos obligatorios vacíos.")
        return redirect(url_for('Crear_contactos'))

    # Separo el nombre completo en nombre y apellido
    # Esto es porque en mi Excel tengo columnas separadas
    partes = nombre_completo.split(' ', 1)
    nombre = partes[0]
    apellido = partes[1] if len(partes) > 1 else ''

    # -------------------- VALIDACIÓN 2: TELÉFONO --------------------
    # En Costa Rica usamos 8 dígitos exactos, todos numéricos
    if not telefono.isdigit() or len(telefono) != 8:
        print("Error de Backend: El teléfono debe poseer 8 dígitos numéricos.")
        return redirect(url_for('Crear_contactos'))

    # -------------------- VALIDACIÓN 3: CORREO (Indicador 5) --------------------
    # Uso la expresión regular para asegurar el formato del correo
    if not re.match(EMAIL_REGEX, correo):
        print("Error de Backend: Formato de correo electrónico inválido.")
        return redirect(url_for('Crear_contactos'))

    # -------------------- GUARDADO EN EXCEL --------------------
    # Si todas las validaciones pasaron, guardo el contacto
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    # Agrego una nueva fila con los datos, el favorito inicia en "No"
    ws.append([nombre, apellido, telefono, correo, direccion, categoria, "No"])
    wb.save(EXCEL_FILE)

    print(f"Éxito: Contacto [{nombre} {apellido}] almacenado en Excel.")
    return redirect(url_for('principal'))

# =========================================================================
# RUTA PARA MARCAR/QUITAR FAVORITOS
# =========================================================================

@app.route('/toggle_favorito/<int:id_fila>')
def toggle_favorito(id_fila):
    """
    Esta función cambia el estado de favorito de un contacto.
    Si estaba en 'Sí' lo pasa a 'No', y viceversa.
    Es como un interruptor que enciende y apaga el favorito.
    """
    # Cargo el archivo Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Verifico que el ID sea válido (que la fila exista)
    if 2 <= id_fila <= ws.max_row:
        # Obtengo el valor actual de favorito
        valor_actual = ws.cell(row=id_fila, column=7).value
        
        # Cambio el valor (Sí -> No, No -> Sí)
        if valor_actual == 'Sí':
            ws.cell(row=id_fila, column=7).value = 'No'
        else:
            ws.cell(row=id_fila, column=7).value = 'Sí'
            
        wb.save(EXCEL_FILE)  # Guardo los cambios
    
    # Redirijo a la página de origen (para que el usuario no se pierda)
    origen = request.args.get('origen', 'principal')
    if origen == 'mas_info':
        return redirect(url_for('mas_informacion', id_fila=id_fila))
    return redirect(url_for('principal'))

# =========================================================================
# RUTA PARA VER INFORMACIÓN DETALLADA DE UN CONTACTO
# =========================================================================

@app.route('/mas_informacion/<int:id_fila>')
def mas_informacion(id_fila):
    """
    Esta función muestra el reporte detallado de un contacto específico.
    Es como la ficha personal de cada contacto con toda su información.
    """
    # Cargo el Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    contacto = None
    
    # Busco el contacto por su ID (número de fila)
    if 2 <= id_fila <= ws.max_row:
        nom = ws.cell(row=id_fila, column=1).value
        if nom:  # Si existe el nombre, armo el diccionario completo
            contacto = {
                'id_fila': id_fila,
                'nombre': nom,
                'apellido': ws.cell(row=id_fila, column=2).value,
                'telefono': ws.cell(row=id_fila, column=3).value,
                'correo': ws.cell(row=id_fila, column=4).value,
                'direccion': ws.cell(row=id_fila, column=5).value,
                'categoria': ws.cell(row=id_fila, column=6).value,
                'favorito': ws.cell(row=id_fila, column=7).value,
                'estado': 'Conectado'  # Un detalle extra para la vista
            }
    
    # Si encontré el contacto, muestro su información detallada
    if contacto:
        return render_template('mas_informacion.html', contacto=contacto, id_fila=id_fila)
    return redirect(url_for('principal'))

# =========================================================================
# RUTA PARA ELIMINAR CONTACTOS
# =========================================================================

@app.route('/eliminar_contacto/<int:id_fila>')
def eliminar_contacto(id_fila):
    """
    Esta función elimina físicamente un contacto del archivo Excel.
    Borra toda la fila correspondiente al ID que le pasen.
    ¡Cuidado! Esta acción no se puede deshacer.
    """
    # Cargo el Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Verifico que el ID sea válido
    if 2 <= id_fila <= ws.max_row:
        # Elimino la fila completa
        ws.delete_rows(id_fila, 1)
        wb.save(EXCEL_FILE)  # Guardo los cambios
        print(f"Éxito: Registro de la fila {id_fila} eliminado del archivo Excel.")
    
    return redirect(url_for('principal'))

# =========================================================================
# RUTA PARA EDITAR CONTACTOS - FORMULARIO
# =========================================================================

@app.route('/editar/<int:id_fila>', methods=['GET'])
def editar_usuario(id_fila):
    """
    Esta función carga los datos de un contacto en el formulario de edición.
    Es como "traer" la información para que el usuario pueda modificarla.
    """
    # Cargo el Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    contacto = None
    
    # Busco el contacto por su ID
    if 2 <= id_fila <= ws.max_row:
        nom = ws.cell(row=id_fila, column=1).value
        ape = ws.cell(row=id_fila, column=2).value
        
        if nom:
            # Armo el nombre completo para mostrarlo en el formulario
            nombre_completo = f"{nom} {ape}".strip() if ape else nom
            contacto = {
                'nombre_completo': nombre_completo,
                'telefono': ws.cell(row=id_fila, column=3).value,
                'correo': ws.cell(row=id_fila, column=4).value,
                'direccion': ws.cell(row=id_fila, column=5).value,
                'categoria': ws.cell(row=id_fila, column=6).value
            }
    
    # Si encontré el contacto, muestro el formulario de edición
    if contacto:
        return render_template('Editar.html', contacto=contacto, id_fila=id_fila)
    return redirect(url_for('principal'))

# =========================================================================
# RUTA PARA ACTUALIZAR CONTACTOS - PROCESAMIENTO
# =========================================================================

@app.route('/actualizar_contacto/<int:id_fila>', methods=['POST'])
def actualizar_contacto(id_fila):
    """
    Esta función procesa la edición de un contacto. Sobrescribe los datos
    existentes con los nuevos que envió el usuario, pero antes pasa por
    las mismas validaciones que al crear un contacto nuevo.
    """
    # Obtengo todos los datos del formulario de edición
    nombre_completo = request.form.get('nombre', '').strip()
    telefono_nuevo = request.form.get('telefono', '').strip()
    correo = request.form.get('correo', '').strip()
    categoria = request.form.get('categoria', 'Otro')
    direccion = request.form.get('direccion', '').strip()

    # -------------------- VALIDACIONES DE BACKEND EN EDICIÓN --------------------
    # Verifico que los campos obligatorios no estén vacíos
    if not nombre_completo or not telefono_nuevo or not correo:
        print("Error de Backend: Campos vacíos detectados en proceso de edición.")
        return redirect(url_for('editar_usuario', id_fila=id_fila))

    # Separo nombre y apellido como siempre
    partes = nombre_completo.split(' ', 1)
    nombre = partes[0]
    apellido = partes[1] if len(partes) > 1 else ''

    # Valido el teléfono (8 dígitos numéricos)
    if not telefono_nuevo.isdigit() or len(telefono_nuevo) != 8:
        print("Error de Backend: Teléfono editado inválido.")
        return redirect(url_for('editar_usuario', id_fila=id_fila))

    # Valido el correo con la expresión regular
    if not re.match(EMAIL_REGEX, correo):
        print("Error de Backend: Correo editado inválido.")
        return redirect(url_for('editar_usuario', id_fila=id_fila))

    # -------------------- ACTUALIZACIÓN EN EXCEL --------------------
    # Si todo está bien, actualizo los datos del contacto
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Sobrescribo cada campo en su columna correspondiente
    ws.cell(row=id_fila, column=1).value = nombre
    ws.cell(row=id_fila, column=2).value = apellido
    ws.cell(row=id_fila, column=3).value = telefono_nuevo
    ws.cell(row=id_fila, column=4).value = correo
    ws.cell(row=id_fila, column=5).value = direccion
    ws.cell(row=id_fila, column=6).value = categoria
               
    wb.save(EXCEL_FILE)  # Guardo los cambios
    print(f"Éxito: Registro {id_fila} actualizado correctamente en la base de datos.")
    
    # Redirijo a la vista de información detallada del contacto actualizado
    return redirect(url_for('mas_informacion', id_fila=id_fila))

# =========================================================================
# RUTA PARA ESTADÍSTICAS - REPORTES
# =========================================================================

@app.route('/estadisticas')
def estadisticas():
    """
    Esta función genera un reporte estadístico de todos los contactos.
    Cuenta cuántos contactos hay en total, cuántos son favoritos, y
    muestra los contactos más recientes. Es como un mini dashboard.
    """
    # Me aseguro que el Excel exista y lo cargo
    inicializar_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Inicializo los contadores
    total_contactos = 0
    total_favoritos = 0
    lista_recientes = []  # Aquí guardo los contactos para mostrarlos
    
    # Recorro todas las filas (saltándome los encabezados)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Si hay un nombre en la fila
            total_contactos += 1  # Cuento un contacto más
            
            # Guardo los datos para la lista de recientes
            lista_recientes.append({
                'nombre': row[0], 
                'apellido': row[1] if row[1] else '', 
                'correo': row[3]
            })
            
            # Verifico si es favorito (comparación exacta de string)
            if row[6] == 'Sí':
                total_favoritos += 1  # Cuento un favorito más
    
    # La lista_recientes la muestro en orden inverso (los más nuevos primero)
    return render_template('estadistica.html', 
                          total=total_contactos, 
                          favorites=total_favoritos, 
                          contactos_recientes=lista_recientes[::-1])

# =========================================================================
# PUNTO DE ENTRADA DE LA APLICACIÓN
# =========================================================================

if __name__ == '__main__':
    app.run(debug=True)  # Inicio el servidor con modo debug para desarrollo